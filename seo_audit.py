#!/usr/bin/env python3
"""
SEO audit tool for passot.co.jp (WordPress + Elementor, EN at /en/, JP at root).

Discovers pages via sitemap.xml, crawls each one, extracts on-page SEO
signals, flags issues, compares EN/JP counterparts, and writes everything
to a single XLSX workbook.

Usage:
    pip install -r requirements.txt
    python seo_audit.py

Run `python seo_audit.py --help` for all options.
"""

import argparse
import csv
import json
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import dataclass, field
from typing import List, Literal
from urllib.parse import urljoin, urlparse, urldefrag

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

# --------------------------------------------------------------------------
# Config defaults (override via CLI flags — see --help)
# --------------------------------------------------------------------------

DEFAULT_BASE_URL = "https://www.passot.co.jp"
DEFAULT_SITEMAP_URL = "https://www.passot.co.jp/sitemap_index.xml"
DEFAULT_EN_PREFIX = "/en/"
DEFAULT_OUTPUT = "passot_seo_audit.xlsx"
DEFAULT_DELAY = 0.8
DEFAULT_TIMEOUT = 15
DEFAULT_THIN_WORDS_EN = 300
DEFAULT_THIN_CHARS_JA = 600
# Cap on stored main-content text per page — bounds memory and, when
# --ai-analysis is used, bounds per-page token cost. Generous relative to
# observed page sizes (even long-form articles run well under this).
MAX_CONTENT_TEXT_CHARS = 8000
DEFAULT_AI_MODEL = "claude-sonnet-5"
DEFAULT_SEARCH_DATA_DAYS = 28
GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/webmasters.readonly",
    "https://www.googleapis.com/auth/analytics.readonly",
]
DEFAULT_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36 PassotSEOAudit/1.0"
)
# Sub-sitemaps whose filenames contain any of these are skipped (archive
# listings, not real content pages).
DEFAULT_SITEMAP_EXCLUDE = ["author-sitemap", "category-sitemap", "tag-sitemap", "feed"]
NON_HTML_EXTENSIONS = {
    ".jpg", ".jpeg", ".png", ".gif", ".svg", ".webp", ".ico",
    ".pdf", ".zip", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".mp4", ".mp3", ".css", ".js", ".xml", ".json",
}
MAIN_CONTENT_SELECTORS = ["main", "article", ".elementor", "#content", ".entry-content", ".site-content"]
BOILERPLATE_SELECTOR = "nav, header, footer, aside, script, style, form"

SITEMAP_NS = {"sm": "http://www.sitemaps.org/schemas/sitemap/0.9"}

HIGH, MEDIUM, LOW, INFO = "High", "Medium", "Low", "Info"
SEVERITY_ORDER = {HIGH: 0, MEDIUM: 1, LOW: 2, INFO: 3}


# --------------------------------------------------------------------------
# Data model
# --------------------------------------------------------------------------

@dataclass
class Page:
    url: str
    lang: str
    sitemap_lastmod: str = ""
    status_code: int = None
    final_url: str = ""
    redirected: bool = False
    fetch_error: str = ""
    title: str = ""
    meta_description: str = ""
    canonical: str = ""
    hreflang: dict = field(default_factory=dict)
    h1_list: list = field(default_factory=list)
    h2_list: list = field(default_factory=list)
    image_total: int = 0
    image_missing_alt: int = 0
    missing_alt_examples: list = field(default_factory=list)
    internal_link_count: int = 0
    internal_links: set = field(default_factory=set)
    all_internal_links: set = field(default_factory=set)
    word_count: int = None
    char_count: int = None
    http_last_modified: str = ""
    structured_data_types: list = field(default_factory=list)
    og_tags: dict = field(default_factory=dict)
    twitter_card: bool = False
    content_text: str = ""
    gsc_clicks: int = None
    gsc_impressions: int = None
    gsc_ctr: float = None
    gsc_avg_position: float = None
    gsc_top_query: str = ""
    ga4_sessions: int = None
    ga4_pageviews: int = None
    ga4_engagement_rate: float = None
    ga4_bounce_rate: float = None
    ga4_conversions: float = None
    external_backlinks: int = None

    @property
    def ok(self):
        return not self.fetch_error and self.status_code == 200


@dataclass
class Issue:
    severity: str
    category: str
    url: str
    lang: str
    detail: str


# --------------------------------------------------------------------------
# Small helpers
# --------------------------------------------------------------------------

def normalize_netloc(netloc):
    netloc = netloc.lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    return netloc


def is_probably_html(url):
    path = urlparse(url).path
    ext = path.rsplit(".", 1)
    if len(ext) == 2 and f".{ext[1].lower()}" in NON_HTML_EXTENSIONS:
        return False
    return True


def normalize_link(href, page_url):
    href = (href or "").strip()
    if not href or href.startswith(("mailto:", "tel:", "javascript:", "#")):
        return None
    absolute = urljoin(page_url, href)
    absolute, _frag = urldefrag(absolute)
    return absolute


def resolve_links(link_tags, page_url, base_netloc):
    result = set()
    for a in link_tags:
        norm = normalize_link(a.get("href", ""), page_url)
        if not norm:
            continue
        if normalize_netloc(urlparse(norm).netloc) != base_netloc:
            continue
        if norm.rstrip("/") == page_url.rstrip("/"):
            continue
        result.add(norm)
    return result


def truncate_join(items, limit=5, sep="; "):
    items = [str(i) for i in items]
    if len(items) <= limit:
        return sep.join(items)
    return sep.join(items[:limit]) + f" ... (+{len(items) - limit} more)"


def classify_lang(url, en_prefix):
    path = urlparse(url).path
    return "en" if path.startswith(en_prefix) else "ja"


# --------------------------------------------------------------------------
# Sitemap discovery
# --------------------------------------------------------------------------

def fetch_xml(session, url, timeout):
    resp = session.get(url, timeout=timeout)
    resp.raise_for_status()
    return ET.fromstring(resp.content)


def collect_sitemap_urls(session, sitemap_url, timeout, exclude_patterns, seen=None):
    if seen is None:
        seen = set()
    if sitemap_url in seen:
        return []
    seen.add(sitemap_url)

    root = fetch_xml(session, sitemap_url, timeout)
    tag = root.tag.lower()
    results = []

    if tag.endswith("sitemapindex"):
        for sm in root.findall("sm:sitemap", SITEMAP_NS):
            loc = (sm.findtext("sm:loc", default="", namespaces=SITEMAP_NS) or "").strip()
            if not loc or any(p in loc.lower() for p in exclude_patterns):
                continue
            results.extend(collect_sitemap_urls(session, loc, timeout, exclude_patterns, seen))
    elif tag.endswith("urlset"):
        for u in root.findall("sm:url", SITEMAP_NS):
            loc = (u.findtext("sm:loc", default="", namespaces=SITEMAP_NS) or "").strip()
            lastmod = (u.findtext("sm:lastmod", default="", namespaces=SITEMAP_NS) or "").strip()
            if loc:
                results.append((loc, lastmod))

    return results


# --------------------------------------------------------------------------
# Page fetching / extraction
# --------------------------------------------------------------------------

def fetch_page(session, url, timeout, retries=3, backoff=2):
    last_exc = None
    for attempt in range(retries):
        try:
            return session.get(url, timeout=timeout, allow_redirects=True), None
        except requests.RequestException as e:
            last_exc = e
            time.sleep(backoff * (attempt + 1))
    return None, str(last_exc)


def check_link_status(session, url, timeout):
    try:
        resp = session.head(url, timeout=timeout, allow_redirects=True)
        if resp.status_code >= 400:
            resp = session.get(url, timeout=timeout, allow_redirects=True, stream=True)
            resp.close()
        return resp.status_code
    except requests.RequestException as e:
        return f"ERROR: {e.__class__.__name__}"


def find_main_container_labeled(soup):
    """Among all candidate containers found on the page, pick the one with
    the most text — not just the first matching selector. Themes/page
    builders (Elementor especially) sometimes render a near-empty <main>
    wrapper alongside the real content in a sibling container (e.g.
    `.elementor`); picking the first match by priority alone can grab the
    empty wrapper and report the page as having ~0 words."""
    candidates = [(selector, el) for selector in MAIN_CONTENT_SELECTORS for el in soup.select(selector)]
    if not candidates:
        return "body-fallback (no selector matched)", (soup.body or soup)
    return max(candidates, key=lambda pair: len(pair[1].get_text(strip=True)))


def find_main_container(soup):
    _, el = find_main_container_labeled(soup)
    return el


def extract_head_fields(soup):
    title = soup.title.get_text(strip=True) if soup.title else ""

    meta_tag = soup.find("meta", attrs={"name": re.compile("^description$", re.I)})
    meta_description = meta_tag["content"].strip() if meta_tag and meta_tag.get("content") else ""

    canonical_tag = soup.find("link", rel=lambda v: v and "canonical" in v)
    canonical = canonical_tag["href"].strip() if canonical_tag and canonical_tag.get("href") else ""

    hreflang = {}
    for link in soup.find_all("link", rel=lambda v: v and "alternate" in v):
        hl, href = link.get("hreflang"), link.get("href")
        if hl and href:
            hreflang[hl.lower()] = href.strip()

    h1_list = [h.get_text(strip=True) for h in soup.find_all("h1")]
    h2_list = [h.get_text(strip=True) for h in soup.find_all("h2")]
    return title, meta_description, canonical, hreflang, h1_list, h2_list


def extract_structured_data(soup):
    """Presence-only check: JSON-LD @type values, Open Graph tags, Twitter Card.
    Does not validate schema correctness, just whether it's there at all."""
    ld_types = []
    for script in soup.find_all("script", type="application/ld+json"):
        try:
            data = json.loads(script.string or "")
        except (ValueError, TypeError):
            continue
        for item in data if isinstance(data, list) else [data]:
            if not isinstance(item, dict):
                continue
            for node in item.get("@graph", [item]):
                if not isinstance(node, dict):
                    continue
                t = node.get("@type")
                if isinstance(t, list):
                    ld_types.extend(str(x) for x in t)
                elif t:
                    ld_types.append(str(t))

    og_tags = {}
    for meta in soup.find_all("meta", property=re.compile("^og:", re.I)):
        content = (meta.get("content") or "").strip()
        if content:
            og_tags[meta["property"].lower()] = content

    twitter_card = bool(soup.find("meta", attrs={"name": re.compile("^twitter:card$", re.I)}))

    return sorted(set(ld_types)), og_tags, twitter_card


def extract_content_metrics(soup, lang):
    container = find_main_container(soup)
    # Re-parse a copy so we can strip boilerplate without mutating `soup`
    # (the caller still needs the untouched tree for full-page link scanning).
    working = BeautifulSoup(str(container), "lxml")
    for tag in working.select(BOILERPLATE_SELECTOR):
        tag.decompose()

    text = working.get_text(separator=" ", strip=True)
    images = working.find_all("img")
    missing_alt = [img.get("src", "") for img in images if not img.get("alt", "").strip()]

    word_count = char_count = None
    if lang == "en":
        word_count = len(text.split())
    else:
        char_count = len(re.sub(r"\s+", "", text))

    main_link_tags = working.find_all("a", href=True)
    return word_count, char_count, len(images), missing_alt, main_link_tags, text[:MAX_CONTENT_TEXT_CHARS]


def build_page(session, url, lang, sitemap_lastmod, timeout, base_netloc):
    page = Page(url=url, lang=lang, sitemap_lastmod=sitemap_lastmod)
    resp, err = fetch_page(session, url, timeout)
    if err or resp is None:
        page.fetch_error = err or "Unknown error"
        return page

    page.status_code = resp.status_code
    page.final_url = resp.url
    page.redirected = bool(resp.history) and resp.url.rstrip("/") != url.rstrip("/")
    page.http_last_modified = resp.headers.get("Last-Modified", "")

    if resp.status_code != 200 or "text/html" not in resp.headers.get("Content-Type", ""):
        return page

    soup = BeautifulSoup(resp.text, "lxml")

    page.all_internal_links = resolve_links(soup.find_all("a", href=True), url, base_netloc)
    (page.title, page.meta_description, page.canonical, page.hreflang,
     page.h1_list, page.h2_list) = extract_head_fields(soup)
    (page.structured_data_types, page.og_tags, page.twitter_card) = extract_structured_data(soup)
    (page.word_count, page.char_count, page.image_total,
     page.missing_alt_examples, main_link_tags, page.content_text) = extract_content_metrics(soup, lang)
    page.image_missing_alt = len(page.missing_alt_examples)
    page.internal_links = resolve_links(main_link_tags, url, base_netloc)
    page.internal_link_count = len(page.internal_links)
    return page


# --------------------------------------------------------------------------
# Cross-page analysis
# --------------------------------------------------------------------------

def find_duplicates(pages, field_getter):
    groups = defaultdict(list)
    for p in pages:
        if not p.ok:
            continue
        val = field_getter(p)
        if not val:
            continue
        groups[(p.lang, val.strip().lower())].append(p.url)
    return {k: v for k, v in groups.items() if len(v) > 1}


def pair_pages(pages, en_prefix):
    by_url = {p.url.rstrip("/"): p for p in pages}
    en_pages = [p for p in pages if p.lang == "en" and p.ok]
    pairs, matched_en, matched_ja = [], set(), set()

    for en_page in en_pages:
        candidate = None
        for hl, href in en_page.hreflang.items():
            if hl.startswith("ja"):
                candidate = by_url.get(href.rstrip("/"))
                break
        if candidate is None:
            parsed = urlparse(en_page.url)
            path = parsed.path
            if path.startswith(en_prefix):
                ja_path = "/" + path[len(en_prefix):]
            elif path.rstrip("/") == en_prefix.rstrip("/"):
                ja_path = "/"
            else:
                ja_path = None
            if ja_path is not None:
                guess = f"{parsed.scheme}://{parsed.netloc}{ja_path}".rstrip("/")
                candidate = by_url.get(guess)
        if candidate:
            pairs.append((en_page, candidate))
            matched_en.add(en_page.url)
            matched_ja.add(candidate.url)

    unmatched_en = [p for p in en_pages if p.url not in matched_en]
    unmatched_ja = [p for p in pages if p.lang == "ja" and p.ok and p.url not in matched_ja]
    return pairs, unmatched_en, unmatched_ja


# Thin-content thresholds are not a real Google ranking signal (there is no
# official minimum word count) — they're a heuristic proxy for "might not be
# substantive," so one number for an entire site is a blunt instrument. A
# product spec page and a competitive blog article don't need the same bar.
# EXEMPT means "this page type is legitimately short by design, don't flag it
# at all" (e.g. a homepage or a blog listing page).
EXEMPT = "exempt"

# (label, word threshold for EN, char threshold for JP) keyed by first URL
# path segment (after stripping the EN prefix) or, for Utility, by any
# hyphen-separated token in the path. Tune these to match your own site's
# URL structure and content strategy.
PAGE_TYPE_BY_FIRST_SEGMENT = {
    "products": ("Product", 150, 300),
    "retail-displays": ("Product", 150, 300),
    "company": ("Company/About", 150, 300),
    "about": ("Company/About", 150, 300),
    "our-process": ("Company/About", 150, 300),
    "process-retail-display": ("Company/About", 150, 300),
}
UTILITY_TOKENS = {"contact", "faq", "privacy", "terms", "security", "policy"}
BLOG_INDEX_SEGMENTS = {"column", "blog"}


def classify_page_type_by_path(url, en_prefix):
    path = urlparse(url).path
    if en_prefix and path.startswith(en_prefix):
        path = "/" + path[len(en_prefix):]
    path = path.strip("/")
    if not path:
        return "Homepage", EXEMPT, EXEMPT

    segments = path.split("/")
    if len(segments) == 1 and segments[0] in BLOG_INDEX_SEGMENTS:
        return "Blog Index", EXEMPT, EXEMPT
    if segments[0] in PAGE_TYPE_BY_FIRST_SEGMENT:
        return PAGE_TYPE_BY_FIRST_SEGMENT[segments[0]]

    tokens = set(re.split(r"[-_]", path.lower()))
    if tokens & UTILITY_TOKENS:
        return "Utility", 80, 150
    if segments[0].startswith("column") or "column" in tokens:
        return "Blog/Column Article", 600, 1000

    return None


def classify_page_types(pages, pairs, en_prefix, default_words, default_chars):
    """URL-path classification first. EN pages with no keyword match (e.g.
    SEO-friendly slugs that share no vocabulary with their JP counterpart's
    URL) inherit their JP pair's classification instead of guessing — that's
    real matched data, not a fabricated signal. Anything still unclassified
    falls back to the site-wide --thin-words/--thin-chars default."""
    by_url = {}
    for p in pages:
        by_url[p.url] = classify_page_type_by_path(p.url, en_prefix)

    for en_page, ja_page in pairs:
        if by_url.get(en_page.url) is None and by_url.get(ja_page.url) is not None:
            by_url[en_page.url] = by_url[ja_page.url]

    result = {}
    for p in pages:
        classified = by_url.get(p.url)
        if classified is None:
            result[p.url] = ("Other", default_words, default_chars)
        else:
            result[p.url] = classified
    return result


def check_broken_links(session, pages, timeout, delay, skip=False):
    crawled_status = {p.url.rstrip("/"): ("ERROR" if p.fetch_error else p.status_code) for p in pages}

    link_sources = defaultdict(set)
    for p in pages:
        for link in p.all_internal_links:
            link_sources[link].add(p.url)

    results = {}
    if skip:
        return results, link_sources

    for i, link in enumerate(sorted(link_sources), 1):
        key = link.rstrip("/")
        if key in crawled_status:
            results[link] = crawled_status[key]
        else:
            results[link] = check_link_status(session, link, timeout)
            time.sleep(delay)
    return results, link_sources


def compute_inbound_link_counts(pages):
    """How many other crawled pages link to each page from their main content
    (nav/footer links excluded, since those are identical on every page and
    would mask genuinely under-linked content)."""
    urls_by_key = {p.url.rstrip("/"): p.url for p in pages}
    inbound = defaultdict(set)
    for p in pages:
        if not p.ok:
            continue
        for link in p.internal_links:
            key = link.rstrip("/")
            target = urls_by_key.get(key)
            if target and target != p.url:
                inbound[target].add(p.url)
    return inbound


# --------------------------------------------------------------------------
# Issue generation
# --------------------------------------------------------------------------

# Homepages are expected to have few/no inbound links from other pages'
# main content (they're reached via nav, not content links) — exclude them
# from the "weakly linked" check to avoid noise.
WEAK_INBOUND_THRESHOLD = 1

def generate_issues(pages, pairs, unmatched_en, unmatched_ja, broken_links, link_sources,
                     dup_titles, dup_meta, thin_words, thin_chars, en_prefix=DEFAULT_EN_PREFIX):
    issues = []
    inbound_counts = compute_inbound_link_counts(pages)
    page_types = classify_page_types(pages, pairs, en_prefix, thin_words, thin_chars)
    homepage_paths = {"/", en_prefix.rstrip("/") or "/en"}

    def add(sev, cat, url, lang, detail):
        issues.append(Issue(sev, cat, url, lang, detail))

    for p in pages:
        if p.fetch_error:
            add(HIGH, "Page Unreachable", p.url, p.lang, f"Request failed: {p.fetch_error}")
            continue
        if p.status_code != 200:
            add(HIGH, "Page Unreachable", p.url, p.lang, f"HTTP {p.status_code}")
            continue
        if p.redirected:
            add(LOW, "Sitemap Redirect", p.url, p.lang, f"Sitemap URL redirects to {p.final_url}")

        if not p.title:
            add(HIGH, "Missing Title", p.url, p.lang, "No <title> tag content")

        if not p.meta_description:
            add(MEDIUM, "Missing Meta Description", p.url, p.lang, "No meta description found")

        if not p.canonical:
            add(HIGH, "Missing Canonical", p.url, p.lang, "No canonical link tag found")
        elif p.canonical.rstrip("/") != p.url.rstrip("/") and p.canonical.rstrip("/") != p.final_url.rstrip("/"):
            add(MEDIUM, "Canonical Mismatch", p.url, p.lang, f"Canonical points to {p.canonical}, not this URL")

        if len(p.h1_list) == 0:
            add(MEDIUM, "Missing H1", p.url, p.lang, "No H1 tag found in main content")
        elif len(p.h1_list) > 1:
            add(MEDIUM, "Multiple H1", p.url, p.lang, f"{len(p.h1_list)} H1 tags found: {truncate_join(p.h1_list, 3)}")

        if p.image_missing_alt > 0:
            add(LOW, "Missing Alt Text", p.url, p.lang,
                f"{p.image_missing_alt}/{p.image_total} images missing alt text: "
                f"{truncate_join(p.missing_alt_examples, 3)}")

        page_type, word_th, char_th = page_types.get(p.url, ("Other", thin_words, thin_chars))
        threshold = word_th if p.lang == "en" else char_th
        content_len = p.word_count if p.lang == "en" else p.char_count
        unit = "words" if p.lang == "en" else "characters"
        if threshold != EXEMPT and content_len is not None and content_len < threshold:
            add(MEDIUM, "Thin Content", p.url, p.lang,
                f"{content_len} {unit} of main content (threshold {threshold} for page type '{page_type}')")

        if not p.structured_data_types:
            add(LOW, "Missing Structured Data", p.url, p.lang, "No JSON-LD structured data (schema.org) found on this page")

        missing_og = [k for k in ("og:title", "og:description", "og:image") if k not in p.og_tags]
        if missing_og:
            add(LOW, "Missing Open Graph Tags", p.url, p.lang,
                f"Missing: {', '.join(missing_og)} (affects link preview appearance on social/chat apps)")

        path = urlparse(p.url).path.rstrip("/") or "/"
        if path not in homepage_paths:
            inbound = len(inbound_counts.get(p.url, set()))
            if inbound <= WEAK_INBOUND_THRESHOLD:
                add(LOW, "Weakly Linked Internally", p.url, p.lang,
                    f"Only {inbound} other page(s) link to this page from their main content — "
                    "consider adding internal links from related pages")

    for (lang, _text), urls in dup_titles.items():
        add(MEDIUM, "Duplicate Title", urls[0], lang, f"Same title used on {len(urls)} pages: {truncate_join(urls)}")

    for (lang, _text), urls in dup_meta.items():
        add(MEDIUM, "Duplicate Meta Description", urls[0], lang,
            f"Same meta description used on {len(urls)} pages: {truncate_join(urls)}")

    for link, status in broken_links.items():
        is_broken = isinstance(status, str) or status >= 400
        if is_broken:
            sources = link_sources.get(link, set())
            add(HIGH, "Broken Internal Link", link, "", f"Status: {status}. Linked from: {truncate_join(sources, 4)}")

    for en_page, ja_page in pairs:
        en_to_ja = en_page.hreflang.get("ja") or en_page.hreflang.get("ja-jp")
        ja_to_en = ja_page.hreflang.get("en") or ja_page.hreflang.get("en-us") or ja_page.hreflang.get("en-gb")
        en_ok = bool(en_to_ja) and en_to_ja.rstrip("/") == ja_page.url.rstrip("/")
        ja_ok = bool(ja_to_en) and ja_to_en.rstrip("/") == en_page.url.rstrip("/")
        if not (en_ok and ja_ok):
            add(HIGH, "Missing Hreflang Cross-Link", en_page.url, "en",
                f"hreflang pairing with JP counterpart {ja_page.url} is missing or incomplete "
                f"(EN->JA present: {en_ok}, JA->EN present: {ja_ok})")

        if bool(en_page.meta_description) != bool(ja_page.meta_description):
            add(MEDIUM, "EN/JP Inconsistency", en_page.url, "en",
                f"Meta description present on only one language version (JP: {ja_page.url})")

        if en_page.meta_description and ja_page.meta_description and \
                en_page.meta_description.strip() == ja_page.meta_description.strip():
            add(MEDIUM, "EN/JP Inconsistency", en_page.url, "en",
                f"Meta description is byte-identical between EN and JP (likely untranslated): {ja_page.url}")

        if en_page.title and ja_page.title and en_page.title.strip() == ja_page.title.strip():
            add(MEDIUM, "EN/JP Inconsistency", en_page.url, "en",
                f"Title is byte-identical between EN and JP (likely untranslated): {ja_page.url}")

        if len(en_page.h1_list) != len(ja_page.h1_list):
            add(LOW, "EN/JP Inconsistency", en_page.url, "en",
                f"H1 count differs: EN={len(en_page.h1_list)} vs JP={len(ja_page.h1_list)} ({ja_page.url})")

    for p in unmatched_en:
        add(INFO, "No Language Counterpart", p.url, p.lang, "No matching JP page found (by hreflang or URL pattern)")
    for p in unmatched_ja:
        add(INFO, "No Language Counterpart", p.url, p.lang, "No matching EN page found (by hreflang or URL pattern)")

    issues.sort(key=lambda i: (SEVERITY_ORDER.get(i.severity, 9), i.category, i.url))
    return issues


def find_recurring_missing_alt(pages, min_pages=2):
    """Images missing alt text on multiple pages usually means one shared
    template component (e.g. a CTA icon), not N separate mistakes."""
    src_to_pages = defaultdict(set)
    for p in pages:
        if not p.ok:
            continue
        for src in p.missing_alt_examples:
            src_to_pages[src].add(p.url)
    return {src: urls for src, urls in src_to_pages.items() if len(urls) >= min_pages}


# One-line recommended action + rough effort per issue category, used in the
# Executive Summary tab. Effort is a coarse hint, not an estimate in hours:
# Quick = a single-field fix, Medium = template/many-page fix, Involved = needs
# real content work, Strategic = a judgment call, not a "bug".
CATEGORY_GUIDANCE = {
    "Page Unreachable": ("Fix the broken URL or update/remove links and sitemap entries pointing to it.", "Quick"),
    "Missing Title": ("Add a unique, descriptive <title> tag.", "Quick"),
    "Duplicate Title": ("Differentiate the title tag on each affected page.", "Quick"),
    "Missing Meta Description": ("Write a unique meta description (roughly 50-160 characters).", "Quick"),
    "Duplicate Meta Description": ("Differentiate the meta description on each affected page.", "Quick"),
    "Missing Canonical": ("Add a self-referencing canonical tag.", "Quick"),
    "Canonical Mismatch": ("Confirm the canonical target is correct; fix if it points to the wrong URL.", "Quick"),
    "Missing H1": ("Add a single, descriptive H1 to the page's main content.", "Quick"),
    "Multiple H1": ("Keep one H1; demote the rest to H2/H3.", "Quick"),
    "Missing Alt Text": ("Add descriptive alt text — check the Executive Summary's recurring-image table first.", "Medium"),
    "Thin Content": ("Expand the page with more substantive, unique copy.", "Involved"),
    "Broken Internal Link": ("Fix or remove the link, or update it to the correct URL.", "Quick"),
    "Sitemap Redirect": ("Point the sitemap/internal links at the final URL directly.", "Quick"),
    "Missing Hreflang Cross-Link": ("Add reciprocal hreflang tags linking the EN and JP versions.", "Medium"),
    "EN/JP Inconsistency": ("Review and align content between the EN and JP versions of the page.", "Medium"),
    "No Language Counterpart": ("Confirm this is intentional (e.g. JP-only blog content); translate if not.", "Strategic"),
    "Missing Structured Data": ("Add JSON-LD schema.org markup (Article, FAQPage, Organization, etc. as relevant).", "Involved"),
    "Missing Open Graph Tags": ("Add og:title, og:description and og:image meta tags for social/chat link previews.", "Quick"),
    "Weakly Linked Internally": ("Add internal links to this page from related content elsewhere on the site.", "Quick"),
}


# --------------------------------------------------------------------------
# AI content & keyword analysis (opt-in via --ai-analysis)
# --------------------------------------------------------------------------

# Five levels, worst to best, so severity is visible at a glance instead of
# collapsing everything short of "Good" into one bucket. RATING_SEVERITY
# below (most-urgent-first) drives sorting in the sheet.
RATING_VALUES = Literal["Critical", "Poor", "Needs Improvement", "Good", "Excellent"]
RATING_SEVERITY = {"Critical": 0, "Poor": 1, "Needs Improvement": 2, "Good": 3, "Excellent": 4}


class PageAIAnalysis(BaseModel):
    keyword_or_topic: str
    keyword_is_inferred: bool
    title_assessment: str
    title_rating: RATING_VALUES
    meta_description_assessment: str
    meta_description_rating: RATING_VALUES
    content_assessment: str
    content_rating: RATING_VALUES
    overall_rating: RATING_VALUES
    suggestions: List[str]


AI_ANALYSIS_SYSTEM_PROMPT = (
    "You are an experienced SEO analyst reviewing a single web page's on-page optimization. "
    "The page may be in English or Japanese, but always write your assessments and suggestions "
    "in English — the person reading this report reads English. The one exception is "
    "'keyword_or_topic': keep that in the page's own language (e.g. a Japanese keyword phrase for "
    "a Japanese page), since that's what a real searcher would actually type, and translating it "
    "would defeat the point. Give concise, specific, and actionable judgments grounded in what is "
    "actually on the page — not generic SEO advice a template could produce. Judge keyword/topic "
    "fit, whether the title and meta description are compelling and relevant (not just present), "
    "and whether the content substantively covers the topic a searcher would expect. Use the five "
    "rating levels with real spread, not clustered in the middle: 'Excellent' (nothing meaningful "
    "to improve), 'Good' (solid, only minor polish possible), 'Needs Improvement' (real gaps worth "
    "fixing but not urgent), 'Poor' (significant problems actively hurting this page's SEO), "
    "'Critical' (severely deficient or missing — fix immediately). Be honest: most pages have room "
    "to improve, so reserve 'Excellent' for genuinely strong pages, and make every suggestion "
    "concrete enough that someone could act on it without asking a follow-up question."
)


def load_keyword_map(path):
    """CSV with 'url' and 'keyword' columns, mapping specific pages to a
    site-owner-provided target keyword/topic instead of letting the AI infer
    one from the page's own content."""
    keyword_map = {}
    with open(path, newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            url = (row.get("url") or row.get("URL") or "").strip()
            keyword = (row.get("keyword") or row.get("Keyword") or "").strip()
            if url and keyword:
                keyword_map[url.rstrip("/")] = keyword
    return keyword_map


def parse_page_selectors(inline_text, file_path):
    """Combines --ai-pages (comma/newline separated) and --ai-pages-file
    (one entry per line, '#'-comments and blank lines ignored) into a single
    deduplicated list of raw entry strings (full URLs or bare paths). A bad
    file path raises — callers must not treat that as "no selection" and
    silently fall back to an unrestricted (full-cost) AI run."""
    entries = []
    if inline_text:
        for chunk in re.split(r"[,\n]", inline_text):
            chunk = chunk.strip()
            if chunk:
                entries.append(chunk)
    if file_path:
        with open(file_path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    entries.append(line)
    seen = set()
    deduped = []
    for e in entries:
        if e not in seen:
            seen.add(e)
            deduped.append(e)
    return deduped


def _selector_keys(value):
    """Normalizes a URL or bare path to (full_url_key, path_key), stripping
    query/fragment so pasted tracking params don't break matching."""
    if not value.startswith(("http://", "https://")):
        value = "/" + value.lstrip("/")
    parsed = urlparse(value)
    path_key = parsed.path.rstrip("/") or "/"
    if parsed.scheme:
        full_key = f"{parsed.scheme}://{parsed.netloc}{path_key}"
    else:
        full_key = None
    return full_key, path_key


def match_page_selectors(pages, entries):
    """Matches raw --ai-pages/--ai-pages-file entries (full URLs or bare
    paths) against crawled Page objects. Returns (matched_urls: set of
    canonical p.url values, unmatched_entries: list of entries that matched
    no page)."""
    by_full = {}
    by_path = defaultdict(list)
    for p in pages:
        full_key, path_key = _selector_keys(p.url)
        by_full[full_key] = p.url
        by_path[path_key].append(p.url)

    matched = set()
    unmatched = []
    for entry in entries:
        full_key, path_key = _selector_keys(entry)
        found = []
        if full_key and full_key in by_full:
            found = [by_full[full_key]]
        elif path_key in by_path:
            found = by_path[path_key]
        if found:
            matched.update(found)
        else:
            unmatched.append(entry)
    return matched, unmatched


def analyze_page_with_ai(client, page, model, target_keyword=None, keyword_source=None):
    if target_keyword and keyword_source == "user_specified":
        keyword_line = (
            f"The site owner has specified this page's target keyword/topic as: {target_keyword!r}. "
            "Judge fit against this specific target."
        )
    elif target_keyword and keyword_source == "search_console":
        keyword_line = (
            f"Google Search Console shows this page's top real search query (by impressions) is: "
            f"{target_keyword!r}. Judge fit against this actual query people are typing — not a guess."
        )
    else:
        keyword_line = (
            "No target keyword was specified — infer the page's likely intended topic/keyword from "
            "its own title, headings, and content, and say so."
        )

    performance_line = ""
    if page.gsc_impressions is not None:
        ctr_text = f"{page.gsc_ctr * 100:.1f}%" if page.gsc_ctr is not None else "n/a"
        position_text = f"{page.gsc_avg_position:.1f}" if page.gsc_avg_position is not None else "n/a"
        performance_line = (
            f"\nReal Search Console performance (recent period): {page.gsc_clicks} clicks, "
            f"{page.gsc_impressions} impressions, CTR {ctr_text}, average position {position_text}. "
            "Factor this into your judgment — e.g. decent position with poor CTR usually points at the "
            "title/meta description, not the content.\n"
        )

    user_content = (
        f"URL: {page.url}\n"
        f"Language: {'English' if page.lang == 'en' else 'Japanese'}\n"
        f"{keyword_line}\n"
        f"{performance_line}\n"
        f"Title tag ({len(page.title)} chars): {page.title!r}\n"
        f"Meta description ({len(page.meta_description)} chars): {page.meta_description!r}\n"
        f"H1: {page.h1_list}\n"
        f"H2s: {page.h2_list}\n\n"
        f"Main page content:\n{page.content_text}"
    )

    response = client.messages.parse(
        model=model,
        max_tokens=2048,
        # This is a single-page classification/rating task, not multi-step
        # reasoning — thinking isn't needed, and Sonnet 5 runs it by default
        # when the field is omitted, silently eating into max_tokens and
        # risking the structured JSON output getting cut off mid-response.
        thinking={"type": "disabled"},
        system=AI_ANALYSIS_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": user_content}],
        output_format=PageAIAnalysis,
    )
    if response.parsed_output is None:
        # Structured output didn't parse (e.g. truncated at max_tokens) —
        # some SDK paths return None here instead of raising. Turn it into
        # a real exception so the caller's try/except always catches it,
        # rather than silently storing a None result that crashes later.
        raise ValueError(f"Model returned no parseable output (stop_reason={response.stop_reason!r})")
    return response.parsed_output


def run_ai_analysis(pages, model, keyword_map, selected_urls=None):
    import anthropic

    client = anthropic.Anthropic()
    ok_pages = [p for p in pages if p.ok]
    if selected_urls is not None:
        total_ok = len(ok_pages)
        ok_pages = [p for p in ok_pages if p.url in selected_urls]
        print(f"Restricting AI analysis to {len(ok_pages)} of {total_ok} eligible page(s) "
              f"via --ai-pages/--ai-pages-file.")
    results = {}
    print(f"Running AI content/keyword analysis on {len(ok_pages)} pages using {model} ...")
    for i, p in enumerate(ok_pages, 1):
        explicit_keyword = keyword_map.get(p.url.rstrip("/"))
        if explicit_keyword:
            target_keyword, keyword_source = explicit_keyword, "user_specified"
        elif p.gsc_top_query:
            target_keyword, keyword_source = p.gsc_top_query, "search_console"
        else:
            target_keyword, keyword_source = None, None
        print(f"  [{i}/{len(ok_pages)}] {p.url}")
        try:
            results[p.url] = analyze_page_with_ai(client, p, model, target_keyword, keyword_source)
        except Exception as e:
            print(f"    AI analysis failed: {e}", file=sys.stderr)
    return results


# --------------------------------------------------------------------------
# Google Search Console & GA4 integration (opt-in via --search-data)
# --------------------------------------------------------------------------

def get_google_credentials(credentials_file, scopes):
    from google.oauth2.service_account import Credentials

    if credentials_file:
        return Credentials.from_service_account_file(credentials_file, scopes=scopes)

    raw_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if raw_json:
        return Credentials.from_service_account_info(json.loads(raw_json), scopes=scopes)

    app_creds_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if app_creds_path:
        return Credentials.from_service_account_file(app_creds_path, scopes=scopes)

    raise RuntimeError(
        "No Google credentials found. Pass --google-credentials-file, or set "
        "GOOGLE_SERVICE_ACCOUNT_JSON (the service account JSON key content) or "
        "GOOGLE_APPLICATION_CREDENTIALS (a path to the JSON key file) in the environment."
    )


def fetch_gsc_data(credentials, site_url, days):
    import datetime
    from googleapiclient.discovery import build

    # Search Console data typically has a 2-3 day reporting lag.
    end_date = datetime.date.today() - datetime.timedelta(days=3)
    start_date = end_date - datetime.timedelta(days=days)

    service = build("searchconsole", "v1", credentials=credentials)
    request_body = {
        "startDate": start_date.isoformat(),
        "endDate": end_date.isoformat(),
        "dimensions": ["page", "query"],
        "rowLimit": 25000,
    }
    response = service.searchanalytics().query(siteUrl=site_url, body=request_body).execute()
    rows = response.get("rows", [])

    per_page = defaultdict(lambda: {"clicks": 0, "impressions": 0, "position_weighted_sum": 0.0, "queries": []})
    for row in rows:
        page, query = row["keys"]
        clicks = row.get("clicks", 0)
        impressions = row.get("impressions", 0)
        position = row.get("position", 0.0)
        entry = per_page[page]
        entry["clicks"] += clicks
        entry["impressions"] += impressions
        entry["position_weighted_sum"] += position * impressions
        entry["queries"].append((query, impressions))

    results = {}
    for page, entry in per_page.items():
        impressions = entry["impressions"]
        top_query = max(entry["queries"], key=lambda q: q[1])[0] if entry["queries"] else ""
        results[page] = {
            "clicks": entry["clicks"],
            "impressions": impressions,
            "ctr": (entry["clicks"] / impressions) if impressions else None,
            "avg_position": (entry["position_weighted_sum"] / impressions) if impressions else None,
            "top_query": top_query,
        }
    return results


def _read_gsc_export_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        return [row for row in csv.reader(f)]


def _read_gsc_export_xlsx(path):
    # GSC's Excel export has one sheet per dimension (Queries, Pages,
    # Countries, Devices, ...). Prefer the per-page one.
    wb = load_workbook(path, data_only=True, read_only=True)
    preferred = [n for n in wb.sheetnames if "page" in n.lower()]
    sheet = wb[preferred[0]] if preferred else wb[wb.sheetnames[0]]
    return [["" if v is None else v for v in row] for row in sheet.iter_rows(values_only=True)]


def _find_col(header, *aliases):
    normalized = [str(h).strip().lower() for h in header]
    for alias in aliases:
        if alias in normalized:
            return normalized.index(alias)
    for alias in aliases:
        for i, h in enumerate(normalized):
            if alias in h:
                return i
    return None


def _to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_ctr_fraction(value):
    """CSV exports write CTR as text like "3.5%"; XLSX exports write it as a
    raw fraction cell (0.035). Key off the literal '%' rather than magnitude,
    since a percentage under 1% (e.g. "0.6%") would otherwise be
    indistinguishable from an already-fractional value."""
    if value is None:
        return None
    is_percent_text = isinstance(value, str) and "%" in value
    number = _to_number(value)
    if number is None:
        return None
    return number / 100 if is_percent_text else number


def load_gsc_export(path):
    """Loads a manually-exported Search Console Performance report (CSV or
    Excel, "Pages" tab) as a credential-free alternative to fetch_gsc_data()
    for accounts without Search Console admin access. Returns the same shape:
    {page_url: {clicks, impressions, ctr, avg_position, top_query}}. Per-page
    top query isn't available from this export, so it's left blank."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        rows = _read_gsc_export_xlsx(path)
    elif ext == ".csv":
        rows = _read_gsc_export_csv(path)
    else:
        raise ValueError(f"Unsupported file type for --gsc-import: '{ext}' (use .csv or .xlsx)")

    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        raise ValueError(f"No rows found in {path}")

    header, data_rows = rows[0], rows[1:]
    page_col = _find_col(header, "top pages", "page", "landing page", "url")
    clicks_col = _find_col(header, "clicks")
    impressions_col = _find_col(header, "impressions")
    ctr_col = _find_col(header, "ctr")
    position_col = _find_col(header, "position", "average position", "avg. position", "avg position")

    if page_col is None or clicks_col is None or impressions_col is None:
        raise ValueError(
            f"Could not find Page/Clicks/Impressions columns in {path}. Found headers: {header}"
        )

    results = {}
    for row in data_rows:
        if page_col >= len(row) or not str(row[page_col]).strip():
            continue
        url = str(row[page_col]).strip()
        clicks = _to_number(row[clicks_col]) if clicks_col is not None and clicks_col < len(row) else 0
        impressions = _to_number(row[impressions_col]) if impressions_col is not None and impressions_col < len(row) else 0
        ctr = _to_ctr_fraction(row[ctr_col]) if ctr_col is not None and ctr_col < len(row) else None
        position = _to_number(row[position_col]) if position_col is not None and position_col < len(row) else None
        results[url] = {
            "clicks": int(clicks or 0),
            "impressions": int(impressions or 0),
            "ctr": ctr,
            "avg_position": position,
            "top_query": "",
        }
    return results


def _classify_gsc_links_header(header):
    """Identifies which of Search Console's three "Links" report exports a
    file is (Top linked pages / Top linking sites / Top linking text) from
    its header row, since each is exported as its own file with its own
    column names. Checked most-distinctive-first: a domain/site column is
    unambiguous, an anchor-text column is next, and a page/URL column is
    checked last since URL-shaped values could incidentally appear
    elsewhere. Raises with the header dumped if nothing (or ambiguously
    more than one candidate signal) matches, rather than guessing."""
    site_col = _find_col(header, "linking site", "referring site", "referring domain", "site", "domain")
    text_col = _find_col(header, "linking text", "anchor text", "link text", "text")
    page_col = _find_col(header, "target page", "linked page", "top pages", "page", "url")

    if site_col is not None:
        return "sites"
    if text_col is not None:
        return "text"
    if page_col is not None:
        return "pages"
    raise ValueError(
        f"Could not classify this file as a Search Console Links export "
        f"(no site/domain, anchor-text, or page/URL column found). Found headers: {header}"
    )


def load_gsc_links_export(path):
    """Loads one file from Search Console's "Links" report (External links:
    Top linked pages / Top linking sites / Top linking text), auto-detecting
    which table it is from its header row. Returns (table_type, data):
    table_type is 'pages' | 'sites' | 'text'; data is a {url: count} dict
    for 'pages', or a [(name, count), ...] list for 'sites'/'text'."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        rows = _read_gsc_export_xlsx(path)
    elif ext == ".csv":
        rows = _read_gsc_export_csv(path)
    else:
        raise ValueError(f"Unsupported file type for --gsc-links-import: '{ext}' (use .csv or .xlsx)")

    rows = [r for r in rows if any(str(c).strip() for c in r)]
    if not rows:
        raise ValueError(f"No rows found in {path}")

    header, data_rows = rows[0], rows[1:]
    table_type = _classify_gsc_links_header(header)

    if table_type == "pages":
        name_col = _find_col(header, "target page", "linked page", "top pages", "page", "url")
    elif table_type == "sites":
        name_col = _find_col(header, "linking site", "referring site", "referring domain", "site", "domain")
    else:
        name_col = _find_col(header, "linking text", "anchor text", "link text", "text")
    count_col = _find_col(header, "linking pages", "external links", "links", "count")

    if name_col is None or count_col is None:
        raise ValueError(
            f"Could not find the name/count columns in {path} (classified as '{table_type}'). "
            f"Found headers: {header}"
        )

    if table_type == "pages":
        results = {}
        for row in data_rows:
            if name_col >= len(row) or not str(row[name_col]).strip():
                continue
            name = str(row[name_col]).strip()
            count = _to_number(row[count_col]) if count_col < len(row) else 0
            results[name] = int(count or 0)
        return table_type, results

    results = []
    for row in data_rows:
        if name_col >= len(row) or not str(row[name_col]).strip():
            continue
        name = str(row[name_col]).strip()
        count = _to_number(row[count_col]) if count_col < len(row) else 0
        results.append((name, int(count or 0)))
    return table_type, results


def fetch_ga4_data(credentials, property_id, days):
    from google.analytics.data_v1beta import BetaAnalyticsDataClient
    from google.analytics.data_v1beta.types import DateRange, Dimension, Metric, RunReportRequest

    client = BetaAnalyticsDataClient(credentials=credentials)
    request = RunReportRequest(
        property=f"properties/{property_id}",
        dimensions=[Dimension(name="pagePath")],
        metrics=[
            Metric(name="sessions"),
            Metric(name="screenPageViews"),
            Metric(name="engagementRate"),
            Metric(name="bounceRate"),
            Metric(name="conversions"),
        ],
        date_ranges=[DateRange(start_date=f"{days}daysAgo", end_date="today")],
        limit=100000,
    )
    response = client.run_report(request)

    results = {}
    for row in response.rows:
        path = row.dimension_values[0].value
        sessions, pageviews, engagement_rate, bounce_rate, conversions = (v.value for v in row.metric_values)
        results[path] = {
            "sessions": int(float(sessions)),
            "pageviews": int(float(pageviews)),
            "engagement_rate": float(engagement_rate),
            "bounce_rate": float(bounce_rate),
            "conversions": float(conversions),
        }
    return results


def attach_search_data(pages, gsc_data, ga4_data):
    """Matches fetched GSC/GA4 rows onto crawled Page objects. Returns GSC
    rows that had impressions but matched no crawled page — often orphan
    pages (indexed and getting traffic, but missing from the sitemap)."""
    gsc_by_key = {u.rstrip("/"): d for u, d in (gsc_data or {}).items()}
    ga4_by_path = {(p.rstrip("/") or "/"): d for p, d in (ga4_data or {}).items()}
    matched_gsc_keys = set()

    for p in pages:
        key = p.url.rstrip("/")
        gsc = gsc_by_key.get(key)
        if gsc:
            matched_gsc_keys.add(key)
            p.gsc_clicks = gsc["clicks"]
            p.gsc_impressions = gsc["impressions"]
            p.gsc_ctr = gsc["ctr"]
            p.gsc_avg_position = gsc["avg_position"]
            p.gsc_top_query = gsc["top_query"]

        path_key = urlparse(p.url).path.rstrip("/") or "/"
        ga4 = ga4_by_path.get(path_key)
        if ga4:
            p.ga4_sessions = ga4["sessions"]
            p.ga4_pageviews = ga4["pageviews"]
            p.ga4_engagement_rate = ga4["engagement_rate"]
            p.ga4_bounce_rate = ga4["bounce_rate"]
            p.ga4_conversions = ga4["conversions"]

    orphan_gsc = [(u, d) for u, d in (gsc_data or {}).items() if u.rstrip("/") not in matched_gsc_keys]
    orphan_gsc.sort(key=lambda kv: -kv[1]["impressions"])
    return orphan_gsc


def attach_gsc_external_links(pages, top_linked_pages):
    """Matches a Search Console "Top linked pages" export onto crawled Page
    objects (same p.url.rstrip('/') matching convention as
    attach_search_data). Returns (url, count) pairs that matched no crawled
    page."""
    by_key = {u.rstrip("/"): count for u, count in (top_linked_pages or {}).items()}
    matched_keys = set()
    for p in pages:
        key = p.url.rstrip("/")
        if key in by_key:
            p.external_backlinks = by_key[key]
            matched_keys.add(key)
    return [(u, count) for u, count in (top_linked_pages or {}).items() if u.rstrip("/") not in matched_keys]


# --------------------------------------------------------------------------
# XLSX report
# --------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SEVERITY_FILL = {
    HIGH: PatternFill("solid", fgColor="FFC7CE"),
    MEDIUM: PatternFill("solid", fgColor="FFEB9C"),
    LOW: PatternFill("solid", fgColor="DDEBF7"),
    INFO: PatternFill("solid", fgColor="F2F2F2"),
}
SEVERITY_FONT = {
    HIGH: Font(color="9C0006"),
    MEDIUM: Font(color="9C6500"),
}


def write_sheet(wb, title, headers, rows, col_widths=None):
    ws = wb.create_sheet(title=title[:31])
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(row)
    widths = col_widths or [18] * len(headers)
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    return ws


def write_issues_sheet(wb, issues):
    headers = ["Severity", "Category", "URL", "Language", "Detail"]
    ws = write_sheet(wb, "Issues Summary", headers, [], col_widths=[10, 26, 45, 10, 90])
    for issue in issues:
        ws.append([issue.severity, issue.category, issue.url, issue.lang, issue.detail])
        row_idx = ws.max_row
        fill = SEVERITY_FILL.get(issue.severity)
        font = SEVERITY_FONT.get(issue.severity)
        if fill:
            ws.cell(row=row_idx, column=1).fill = fill
        if font:
            ws.cell(row=row_idx, column=1).font = font
    # Move to front
    wb.move_sheet("Issues Summary", offset=-len(wb.sheetnames))

    counts = defaultdict(int)
    for i in issues:
        counts[i.severity] += 1
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Totals:").font = Font(bold=True)
    for i, sev in enumerate([HIGH, MEDIUM, LOW, INFO]):
        ws.cell(row=summary_row, column=2 + i, value=f"{sev}: {counts.get(sev, 0)}")


def write_executive_summary(wb, pages, issues, pairs, unmatched_en, unmatched_ja, recurring_alt,
                             ai_results=None, orphan_search_pages=None):
    ws = wb.create_sheet("Executive Summary")
    ok_pages = [p for p in pages if p.ok]
    en_count = sum(1 for p in ok_pages if p.lang == "en")
    ja_count = sum(1 for p in ok_pages if p.lang == "ja")

    state = {"row": 1}

    def header(text, size=14):
        c = ws.cell(row=state["row"], column=1, value=text)
        c.font = Font(bold=True, size=size)
        state["row"] += 2

    def subheader(text):
        c = ws.cell(row=state["row"], column=1, value=text)
        c.font = Font(bold=True, size=11, color="305496")
        state["row"] += 1

    def line(text=""):
        ws.cell(row=state["row"], column=1, value=text)
        state["row"] += 1

    def table_header(cols):
        for idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=state["row"], column=idx, value=col_name)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
        state["row"] += 1

    header("Passot SEO Audit — Executive Summary")
    line(f"Pages audited: {len(ok_pages)}  ({en_count} EN, {ja_count} JP)")
    line(f"EN/JP pairs matched: {len(pairs)}  |  Unmatched: {len(unmatched_en)} EN-only, {len(unmatched_ja)} JP-only")
    state["row"] += 1

    counts_by_sev = defaultdict(int)
    counts_by_cat = defaultdict(lambda: defaultdict(int))
    for i in issues:
        counts_by_sev[i.severity] += 1
        counts_by_cat[i.category][i.severity] += 1

    subheader("Issues by severity")
    for sev in [HIGH, MEDIUM, LOW, INFO]:
        line(f"{sev}: {counts_by_sev.get(sev, 0)}")
    state["row"] += 1

    subheader("Issues by category")
    table_header(["Category", "Count", "Worst Severity", "Recommended Action", "Effort"])
    cat_totals = sorted(counts_by_cat.items(), key=lambda kv: (SEVERITY_ORDER.get(min(kv[1], key=lambda s: SEVERITY_ORDER.get(s, 9)), 9), -sum(kv[1].values())))
    for cat, sev_counts in cat_totals:
        total = sum(sev_counts.values())
        worst = min(sev_counts.keys(), key=lambda s: SEVERITY_ORDER.get(s, 9))
        action, effort = CATEGORY_GUIDANCE.get(cat, ("Review the flagged pages.", "Medium"))
        ws.cell(row=state["row"], column=1, value=cat)
        ws.cell(row=state["row"], column=2, value=total)
        ws.cell(row=state["row"], column=3, value=worst)
        ws.cell(row=state["row"], column=4, value=action)
        ws.cell(row=state["row"], column=5, value=effort)
        state["row"] += 1
    state["row"] += 1

    if recurring_alt:
        subheader("Recurring root causes (fix once, resolves many pages)")
        table_header(["Shared Image", "Pages Affected", "Example Pages"])
        for src, urls in sorted(recurring_alt.items(), key=lambda kv: -len(kv[1]))[:20]:
            ws.cell(row=state["row"], column=1, value=src)
            ws.cell(row=state["row"], column=2, value=len(urls))
            ws.cell(row=state["row"], column=3, value=truncate_join(sorted(urls), 3))
            state["row"] += 1
        state["row"] += 1

    content_gaps = unmatched_en + unmatched_ja
    if content_gaps:
        subheader("Translation / content parity gaps (not in the Issues tab's severity ranking)")
        line(f"{len(unmatched_ja)} JP page(s) have no English version. {len(unmatched_en)} EN page(s) have no Japanese version.")
        line("Not necessarily wrong (e.g. intentionally JP-only blog posts) — but a deliberate content decision, not a bug to fix.")
        table_header(["URL", "Language", "Content Length"])
        for p in sorted(content_gaps, key=lambda x: (x.lang, x.url)):
            if p.lang == "en" and p.word_count is not None:
                length = f"{p.word_count} words"
            elif p.char_count is not None:
                length = f"{p.char_count} characters"
            else:
                length = ""
            ws.cell(row=state["row"], column=1, value=p.url)
            ws.cell(row=state["row"], column=2, value=p.lang)
            ws.cell(row=state["row"], column=3, value=length)
            state["row"] += 1
        state["row"] += 1

    ai_results = {url: r for url, r in (ai_results or {}).items() if r is not None}
    if ai_results:
        subheader("AI content & keyword review")
        rating_counts = defaultdict(int)
        urgent_pages = []
        for url, r in ai_results.items():
            rating_counts[r.overall_rating] += 1
            if r.overall_rating in ("Critical", "Poor"):
                urgent_pages.append(url)
        line(
            f"{len(ai_results)} page(s) reviewed. "
            + ", ".join(f"{lvl}: {rating_counts.get(lvl, 0)}" for lvl in RATING_SEVERITY)
            + "."
        )
        if urgent_pages:
            line("Pages rated 'Poor' or 'Critical' overall: " + truncate_join(sorted(urgent_pages), 8))
        line("Full per-page assessments and suggestions are in the AI Content & Keyword Analysis tab "
             "(sorted worst-rated first).")
        state["row"] += 1

    has_search_data = any(p.gsc_impressions is not None or p.ga4_sessions is not None for p in pages)
    top_traffic_pages = []
    if has_search_data:
        subheader("Real traffic & search performance")
        total_clicks = sum(p.gsc_clicks or 0 for p in pages)
        total_impressions = sum(p.gsc_impressions or 0 for p in pages)
        total_sessions = sum(p.ga4_sessions or 0 for p in pages)
        line(f"Site totals (recent period): {total_clicks} clicks / {total_impressions} impressions "
             f"(Search Console), {total_sessions} sessions (GA4).")

        issue_counts = defaultdict(int)
        for i in issues:
            issue_counts[i.url] += 1

        top_traffic_pages = sorted(
            (p for p in pages if p.ok and ((p.ga4_sessions or 0) + (p.gsc_clicks or 0)) > 0),
            key=lambda p: -((p.ga4_sessions or 0) + (p.gsc_clicks or 0)),
        )[:10]
        if top_traffic_pages:
            line("Top pages by real traffic, with open issue counts (fixing issues here matters most):")
            table_header(["URL", "Sessions (GA4)", "Clicks (GSC)", "Open Issues"])
            for p in top_traffic_pages:
                ws.cell(row=state["row"], column=1, value=p.url)
                ws.cell(row=state["row"], column=2, value=p.ga4_sessions or 0)
                ws.cell(row=state["row"], column=3, value=p.gsc_clicks or 0)
                ws.cell(row=state["row"], column=4, value=issue_counts.get(p.url, 0))
                state["row"] += 1
            state["row"] += 1

        if orphan_search_pages:
            top_orphans = orphan_search_pages[:5]
            line(
                f"{len(orphan_search_pages)} URL(s) get real search impressions but weren't found in the "
                "current crawl/sitemap — possible orphan or removed pages worth checking: "
                + truncate_join([u for u, _d in top_orphans], 5)
            )
        line("Full per-page data is in the Search Performance (GSC) and Traffic (GA4) tabs.")
        state["row"] += 1

    subheader("Fix-first priority list")
    high_cats = sorted(c for c, s in counts_by_cat.items() if HIGH in s)
    med_cats = sorted(c for c, s in counts_by_cat.items() if MEDIUM in s and HIGH not in s)
    priorities = []
    if high_cats:
        priorities.append("Address High severity issues first: " + ", ".join(high_cats) + ".")
    if recurring_alt:
        instances = sum(len(urls) for urls in recurring_alt.values())
        priorities.append(
            f"Fix the {len(recurring_alt)} shared/reused image(s) missing alt text once — "
            f"together they account for {instances} page-level flags."
        )
    if med_cats:
        priorities.append("Schedule Medium severity issues next: " + ", ".join(med_cats) + ".")
    if content_gaps:
        priorities.append(
            f"Decide on the {len(content_gaps)} page(s) with no counterpart in the other language — "
            "translate for parity, or confirm it's intentional (see the table above)."
        )
    if ai_results:
        urgent_count = sum(1 for r in ai_results.values() if r.overall_rating in ("Critical", "Poor"))
        if urgent_count:
            priorities.append(f"Review the {urgent_count} page(s) the AI rated 'Poor' or 'Critical' overall first.")
    if top_traffic_pages:
        issue_counts_for_priority = defaultdict(int)
        for i in issues:
            issue_counts_for_priority[i.url] += 1
        flagged_top_traffic = sum(1 for p in top_traffic_pages if issue_counts_for_priority.get(p.url, 0) > 0)
        if flagged_top_traffic:
            verb = "has" if flagged_top_traffic == 1 else "have"
            priorities.append(
                f"{flagged_top_traffic} of your top 10 highest-traffic pages {verb} open issues — "
                "fixing those matters more than the same fix on a low-traffic page (see the table above)."
            )
    if not priorities:
        priorities.append("No High or Medium severity issues found. Review Low/Info items opportunistically.")
    for idx, text in enumerate(priorities, 1):
        line(f"{idx}. {text}")
    line()
    line("Full page-by-page detail behind every number above is in the Issues Summary tab.")

    for idx, width in enumerate([45, 12, 16, 60, 10], 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.move_sheet("Executive Summary", offset=-len(wb.sheetnames))
    return ws


def build_workbook(pages, issues, pairs, unmatched_en, unmatched_ja, broken_links, link_sources,
                    thin_words, thin_chars, en_prefix=DEFAULT_EN_PREFIX, ai_results=None,
                    orphan_search_pages=None, gsc_linking_sites=None, gsc_linking_text=None):
    wb = Workbook()
    wb.remove(wb.active)

    write_issues_sheet(wb, issues)
    recurring_alt = find_recurring_missing_alt(pages)
    write_executive_summary(wb, pages, issues, pairs, unmatched_en, unmatched_ja, recurring_alt,
                             ai_results, orphan_search_pages)

    # All Pages
    page_types = classify_page_types(pages, pairs, en_prefix, thin_words, thin_chars)
    rows = []
    for p in sorted(pages, key=lambda x: (x.lang, x.url)):
        page_type, word_th, char_th = page_types.get(p.url, ("Other", thin_words, thin_chars))
        threshold = word_th if p.lang == "en" else char_th
        rows.append([
            p.url, p.lang, p.status_code or p.fetch_error, p.title, len(p.title),
            p.meta_description, len(p.meta_description), p.canonical,
            len(p.h1_list), len(p.h2_list), p.image_total, p.image_missing_alt,
            p.internal_link_count, p.word_count if p.lang == "en" else "",
            p.char_count if p.lang == "ja" else "", p.http_last_modified, p.sitemap_lastmod,
            len(p.hreflang), page_type, "exempt" if threshold == EXEMPT else threshold,
        ])
    write_sheet(wb, "All Pages", [
        "URL", "Language", "Status", "Title", "Title Length", "Meta Description",
        "Meta Length", "Canonical", "H1 Count", "H2 Count", "Image Count",
        "Images Missing Alt", "Internal Link Count", "Word Count (EN)",
        "Char Count (JP)", "Last-Modified (HTTP)", "Last-Modified (Sitemap)", "Hreflang Tag Count",
        "Page Type", "Thin-Content Threshold Used",
    ], rows, col_widths=[45, 6, 8, 35, 10, 40, 10, 40, 8, 8, 8, 10, 10, 10, 10, 20, 14, 10, 16, 14])

    # Meta Descriptions
    dup_meta = find_duplicates(pages, lambda p: p.meta_description)
    dup_meta_urls = {u for urls in dup_meta.values() for u in urls}
    rows = []
    for p in sorted(pages, key=lambda x: (x.lang, x.url)):
        if not p.ok:
            continue
        rows.append([p.url, p.lang, p.meta_description, len(p.meta_description),
                     "Yes" if not p.meta_description else "No",
                     "Yes" if p.url in dup_meta_urls else "No"])
    write_sheet(wb, "Meta Descriptions", ["URL", "Language", "Meta Description", "Length", "Missing?", "Duplicate?"],
                rows, col_widths=[45, 6, 60, 8, 8, 10])

    # Titles & Headings
    dup_titles = find_duplicates(pages, lambda p: p.title)
    dup_title_urls = {u for urls in dup_titles.values() for u in urls}
    rows = []
    for p in sorted(pages, key=lambda x: (x.lang, x.url)):
        if not p.ok:
            continue
        rows.append([p.url, p.lang, p.title, len(p.title),
                     "Yes" if p.url in dup_title_urls else "No",
                     len(p.h1_list), truncate_join(p.h1_list, 3),
                     len(p.h2_list), truncate_join(p.h2_list, 5)])
    write_sheet(wb, "Titles & Headings",
                ["URL", "Language", "Title", "Title Length", "Duplicate Title?", "H1 Count", "H1 Text", "H2 Count", "H2 Text"],
                rows, col_widths=[45, 6, 35, 10, 12, 8, 40, 8, 50])

    # Images & Alt Text
    rows = []
    for p in sorted(pages, key=lambda x: (x.lang, x.url)):
        if not p.ok:
            continue
        pct = round(100 * p.image_missing_alt / p.image_total, 1) if p.image_total else 0
        rows.append([p.url, p.lang, p.image_total, p.image_missing_alt, pct,
                     truncate_join(p.missing_alt_examples, 4)])
    write_sheet(wb, "Images & Alt Text",
                ["URL", "Language", "Total Images", "Missing Alt", "% Missing", "Example Sources"],
                rows, col_widths=[45, 6, 10, 10, 10, 60])

    # Internal Links (per page)
    inbound_counts = compute_inbound_link_counts(pages)
    rows = []
    for p in sorted(pages, key=lambda x: (x.lang, x.url)):
        if not p.ok:
            continue
        rows.append([p.url, p.lang, p.internal_link_count, len(inbound_counts.get(p.url, set()))])
    write_sheet(wb, "Internal Links",
                ["URL", "Language", "Outbound Links (main content)", "Inbound Links (from other pages' main content)"],
                rows, col_widths=[45, 6, 22, 30])

    # Structured Data & Social Tags
    rows = []
    for p in sorted(pages, key=lambda x: (x.lang, x.url)):
        if not p.ok:
            continue
        rows.append([
            p.url, p.lang, truncate_join(p.structured_data_types, 5) or "(none)",
            "Yes" if p.og_tags else "No",
            "Yes" if "og:title" in p.og_tags else "No",
            "Yes" if "og:description" in p.og_tags else "No",
            "Yes" if "og:image" in p.og_tags else "No",
            "Yes" if p.twitter_card else "No",
        ])
    write_sheet(wb, "Structured Data & Social", [
        "URL", "Language", "JSON-LD Types Found", "Has Open Graph Tags?",
        "OG Title?", "OG Description?", "OG Image?", "Twitter Card?",
    ], rows, col_widths=[45, 6, 35, 16, 10, 14, 10, 12])

    # Broken Links
    rows = []
    for link, status in sorted(broken_links.items()):
        is_broken = isinstance(status, str) or status >= 400
        if is_broken:
            rows.append([link, status, truncate_join(link_sources.get(link, set()), 6)])
    write_sheet(wb, "Broken Links", ["Broken URL", "Status", "Linked From"], rows, col_widths=[50, 12, 80])

    # Canonical Tags
    rows = []
    for p in sorted(pages, key=lambda x: (x.lang, x.url)):
        if not p.ok:
            continue
        self_ref = p.canonical and (p.canonical.rstrip("/") == p.url.rstrip("/") or p.canonical.rstrip("/") == p.final_url.rstrip("/"))
        rows.append([p.url, p.lang, p.canonical or "(missing)", "Yes" if self_ref else "No"])
    write_sheet(wb, "Canonical Tags", ["URL", "Language", "Canonical Value", "Self-Referencing?"],
                rows, col_widths=[45, 6, 45, 15])

    # EN-JP Comparison
    rows = []
    for en_page, ja_page in sorted(pairs, key=lambda pr: pr[0].url):
        en_to_ja = en_page.hreflang.get("ja") or en_page.hreflang.get("ja-jp")
        ja_to_en = ja_page.hreflang.get("en") or ja_page.hreflang.get("en-us") or ja_page.hreflang.get("en-gb")
        reciprocal = bool(en_to_ja) and bool(ja_to_en) and \
            en_to_ja.rstrip("/") == ja_page.url.rstrip("/") and ja_to_en.rstrip("/") == en_page.url.rstrip("/")
        rows.append([
            en_page.url, ja_page.url, "Yes" if reciprocal else "No",
            en_page.title, ja_page.title,
            "Yes" if en_page.meta_description else "No", "Yes" if ja_page.meta_description else "No",
            "Yes" if (en_page.meta_description and ja_page.meta_description and
                      en_page.meta_description.strip() == ja_page.meta_description.strip()) else "No",
            len(en_page.h1_list), len(ja_page.h1_list),
            en_page.canonical or "(missing)", ja_page.canonical or "(missing)",
        ])
    write_sheet(wb, "EN-JP Comparison", [
        "EN URL", "JP URL", "Hreflang Reciprocal?", "EN Title", "JP Title",
        "EN Meta Present?", "JP Meta Present?", "Meta Identical (Untranslated?)",
        "EN H1 Count", "JP H1 Count", "EN Canonical", "JP Canonical",
    ], rows, col_widths=[40, 40, 12, 30, 30, 10, 10, 16, 8, 8, 40, 40])

    unmatched_rows = [[p.url, p.lang] for p in unmatched_en + unmatched_ja]
    if unmatched_rows:
        ws = wb["EN-JP Comparison"]
        start = ws.max_row + 3
        ws.cell(row=start, column=1, value="Pages with no counterpart:").font = Font(bold=True)
        ws.cell(row=start + 1, column=1, value="URL").font = HEADER_FONT
        ws.cell(row=start + 1, column=2, value="Language").font = HEADER_FONT
        for i, (url, lang) in enumerate(unmatched_rows, start + 2):
            ws.cell(row=i, column=1, value=url)
            ws.cell(row=i, column=2, value=lang)

    if ai_results:
        write_ai_analysis_sheet(wb, pages, ai_results)

    if any(p.gsc_impressions is not None for p in pages):
        write_search_performance_sheet(wb, pages)
    if any(p.ga4_sessions is not None for p in pages):
        write_traffic_sheet(wb, pages)
    if any(p.external_backlinks is not None for p in pages):
        write_external_link_pages_sheet(wb, pages)
    if gsc_linking_sites:
        write_external_linking_sites_sheet(wb, gsc_linking_sites)
    if gsc_linking_text:
        write_external_linking_text_sheet(wb, gsc_linking_text)

    return wb


def write_search_performance_sheet(wb, pages):
    rows = []
    for p in sorted(pages, key=lambda x: -(x.gsc_clicks or 0)):
        if not p.ok or p.gsc_impressions is None:
            continue
        rows.append([
            p.url, p.lang, p.gsc_clicks, p.gsc_impressions,
            round(p.gsc_ctr * 100, 2) if p.gsc_ctr is not None else "",
            round(p.gsc_avg_position, 1) if p.gsc_avg_position is not None else "",
            p.gsc_top_query,
        ])
    write_sheet(wb, "Search Performance (GSC)", [
        "URL", "Language", "Clicks", "Impressions", "CTR (%)", "Avg Position", "Top Query",
    ], rows, col_widths=[45, 6, 10, 12, 10, 12, 35])


def write_traffic_sheet(wb, pages):
    rows = []
    for p in sorted(pages, key=lambda x: -(x.ga4_sessions or 0)):
        if not p.ok or p.ga4_sessions is None:
            continue
        rows.append([
            p.url, p.lang, p.ga4_sessions, p.ga4_pageviews,
            round(p.ga4_engagement_rate * 100, 1) if p.ga4_engagement_rate is not None else "",
            round(p.ga4_bounce_rate * 100, 1) if p.ga4_bounce_rate is not None else "",
            p.ga4_conversions,
        ])
    write_sheet(wb, "Traffic (GA4)", [
        "URL", "Language", "Sessions", "Pageviews", "Engagement Rate (%)", "Bounce Rate (%)", "Conversions",
    ], rows, col_widths=[45, 6, 10, 10, 18, 14, 12])


def write_external_link_pages_sheet(wb, pages):
    rows = []
    for p in sorted(pages, key=lambda x: -(x.external_backlinks or 0)):
        if not p.ok or p.external_backlinks is None:
            continue
        rows.append([p.url, p.lang, p.external_backlinks])
    write_sheet(wb, "External Links - Pages (GSC)", [
        "URL", "Language", "External Backlinks (GSC)",
    ], rows, col_widths=[45, 6, 20])


def write_external_linking_sites_sheet(wb, gsc_linking_sites):
    rows = sorted(gsc_linking_sites, key=lambda kv: -kv[1])
    write_sheet(wb, "External Linking Sites (GSC)", [
        "Referring Domain", "Linking Pages",
    ], rows, col_widths=[50, 16])


def write_external_linking_text_sheet(wb, gsc_linking_text):
    rows = sorted(gsc_linking_text, key=lambda kv: -kv[1])
    write_sheet(wb, "External Linking Text (GSC)", [
        "Anchor Text", "Linking Pages",
    ], rows, col_widths=[50, 16])


RATING_FILL = {
    "Critical": PatternFill("solid", fgColor="FFC7CE"),
    "Poor": PatternFill("solid", fgColor="FFD966"),
    "Needs Improvement": PatternFill("solid", fgColor="FFEB9C"),
    "Good": PatternFill("solid", fgColor="C6EFCE"),
    "Excellent": PatternFill("solid", fgColor="A9D18E"),
}


def write_ai_analysis_sheet(wb, pages, ai_results):
    headers = [
        "URL", "Language", "Keyword/Topic", "Keyword Inferred?",
        "Title Rating", "Title Assessment", "Meta Rating", "Meta Assessment",
        "Content Rating", "Content Assessment", "Overall Rating", "Suggestions",
    ]
    ws = write_sheet(wb, "AI Content & Keyword Analysis", headers, [],
                      col_widths=[45, 6, 25, 10, 14, 45, 14, 45, 14, 45, 14, 60])

    rating_cols = {5: "title_rating", 7: "meta_description_rating", 9: "content_rating", 11: "overall_rating"}
    # Worst overall rating first, so the pages needing the most attention are
    # right at the top instead of scattered alphabetically by URL.
    sortable_pages = [p for p in pages if ai_results.get(p.url)]
    sortable_pages.sort(key=lambda p: (RATING_SEVERITY.get(ai_results[p.url].overall_rating, 9), p.lang, p.url))
    for p in sortable_pages:
        r = ai_results.get(p.url)
        if not r:
            continue
        ws.append([
            p.url, p.lang, r.keyword_or_topic, "Yes" if r.keyword_is_inferred else "No",
            r.title_rating, r.title_assessment,
            r.meta_description_rating, r.meta_description_assessment,
            r.content_rating, r.content_assessment,
            r.overall_rating, "; ".join(r.suggestions),
        ])
        row_idx = ws.max_row
        for col_idx, attr in rating_cols.items():
            fill = RATING_FILL.get(getattr(r, attr))
            if fill:
                ws.cell(row=row_idx, column=col_idx).fill = fill


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def parse_args(argv=None):
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--base-url", default=DEFAULT_BASE_URL)
    ap.add_argument("--sitemap-url", default=DEFAULT_SITEMAP_URL)
    ap.add_argument("--en-prefix", default=DEFAULT_EN_PREFIX)
    ap.add_argument("--output", default=DEFAULT_OUTPUT)
    ap.add_argument("--delay", type=float, default=DEFAULT_DELAY, help="Seconds between requests")
    ap.add_argument("--timeout", type=float, default=DEFAULT_TIMEOUT)
    ap.add_argument("--thin-words", type=int, default=DEFAULT_THIN_WORDS_EN, help="Thin-content threshold for EN pages (words)")
    ap.add_argument("--thin-chars", type=int, default=DEFAULT_THIN_CHARS_JA, help="Thin-content threshold for JP pages (characters)")
    ap.add_argument("--user-agent", default=DEFAULT_USER_AGENT)
    ap.add_argument("--max-pages", type=int, default=None, help="Limit number of pages crawled (for a quick test run)")
    ap.add_argument("--skip-broken-links", action="store_true", help="Skip checking internal links for broken status")
    ap.add_argument("--exclude-sitemap", nargs="*", default=DEFAULT_SITEMAP_EXCLUDE,
                     help="Skip sub-sitemaps whose filename contains any of these substrings")
    ap.add_argument("--debug-url", default=None,
                     help="Fetch a single URL and print diagnostics about what content was extracted "
                          "and why (use this to troubleshoot an unexpected word/char count) instead "
                          "of running a full audit")
    ap.add_argument("--ai-analysis", action="store_true",
                     help="Use the Claude API to judge keyword fit, meta quality, and content quality "
                          "per page (costs a small amount per run — see README). Requires "
                          "ANTHROPIC_API_KEY in the environment.")
    ap.add_argument("--ai-model", default=DEFAULT_AI_MODEL,
                     help=f"Claude model to use for --ai-analysis (default: {DEFAULT_AI_MODEL})")
    ap.add_argument("--ai-keyword-map", default=None,
                     help="Optional CSV file with 'url' and 'keyword' columns giving explicit target "
                          "keywords for specific pages; pages not listed fall back to AI-inferred topic")
    ap.add_argument("--ai-pages", default=None,
                     help="Restrict --ai-analysis (the paid step) to these pages only — comma and/or "
                          "newline separated full URLs or paths. No effect without --ai-analysis. "
                          "Union with --ai-pages-file if both are given.")
    ap.add_argument("--ai-pages-file", default=None,
                     help="Text file, one URL or path per line ('#'-comments and blank lines ignored), "
                          "restricting --ai-analysis to those pages. Union with --ai-pages.")
    ap.add_argument("--search-data", action="store_true",
                     help="Pull real per-page data from Search Console and/or GA4 (needs a Google "
                          "service account — see README). Also grounds --ai-analysis in real search "
                          "queries and performance instead of inferring.")
    ap.add_argument("--gsc-site-url", default=None,
                     help="Verified Search Console property, e.g. https://www.passot.co.jp/ or "
                          "sc-domain:passot.co.jp")
    ap.add_argument("--gsc-import", default=None,
                     help="Path to a Search Console Performance report exported manually (CSV or "
                          "XLSX, 'Pages' tab) — an alternative to --gsc-site-url for accounts without "
                          "Search Console admin access. Works standalone or alongside --search-data "
                          "--ga4-property-id. Doesn't require Google credentials or --search-data.")
    ap.add_argument("--ga4-property-id", default=None,
                     help="GA4 numeric Property ID (Admin > Property Settings)")
    ap.add_argument("--google-credentials-file", default=None,
                     help="Path to the Google service account JSON key file. If omitted, falls back to "
                          "the GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_APPLICATION_CREDENTIALS env vars.")
    ap.add_argument("--search-data-days", type=int, default=DEFAULT_SEARCH_DATA_DAYS,
                     help=f"Lookback window in days for Search Console/GA4 data (default: {DEFAULT_SEARCH_DATA_DAYS})")
    ap.add_argument("--gsc-links-import", nargs="+", default=None, metavar="FILE",
                     help="One to three Search Console 'Links' report exports (Top linked pages / "
                          "Top linking sites / Top linking text) — file type auto-detected from the "
                          "header row. CSV or XLSX. Independent of --gsc-import/--search-data. "
                          "Reports raw counts only, no authority/quality score is invented.")
    return ap.parse_args(argv)


def debug_single_url(session, url, en_prefix, timeout):
    lang = classify_lang(url, en_prefix)
    print(f"Fetching {url}  (classified as: {lang})\n")
    resp, err = fetch_page(session, url, timeout)
    if err or resp is None:
        print(f"FETCH FAILED: {err}")
        return 1

    print(f"HTTP status: {resp.status_code}")
    print(f"Content-Type: {resp.headers.get('Content-Type')}")
    if resp.url != url:
        print(f"Redirected to: {resp.url}")
    if resp.status_code != 200 or "text/html" not in resp.headers.get("Content-Type", ""):
        print("\nNot a 200 HTML response, so no content was extracted from it during a real audit run.")
        return 1

    soup = BeautifulSoup(resp.text, "lxml")
    print(f"Raw HTML size: {len(resp.text)} characters\n")

    print("Main-content selectors checked (in priority order), and what each finds on this page:")
    for selector in MAIN_CONTENT_SELECTORS:
        matches = soup.select(selector)
        if matches:
            lengths = [len(m.get_text(strip=True)) for m in matches]
            print(f"  {selector!r:12} -> {len(matches)} match(es), text length(s): {lengths}")
        else:
            print(f"  {selector!r:12} -> no match")

    chosen_selector, container = find_main_container_labeled(soup)
    print(f"\n=> This audit run would use: {chosen_selector!r}")

    word_count, char_count, image_total, missing_alt, _ = extract_content_metrics(soup, lang)
    print(f"\nAfter stripping {BOILERPLATE_SELECTOR!r} from that container:")
    print(f"  Word count (EN): {word_count}")
    print(f"  Char count (JP): {char_count}")
    print(f"  Images found: {image_total} (missing alt: {len(missing_alt)})")

    working = BeautifulSoup(str(container), "lxml")
    for tag in working.select(BOILERPLATE_SELECTOR):
        tag.decompose()
    preview = working.get_text(separator=" ", strip=True)[:400]
    print(f"\nExtracted text preview (first 400 chars): {preview!r}")

    full_body_len = len(soup.get_text(separator=" ", strip=True))
    extracted_len = (word_count or 0) + (char_count or 0)
    print(f"\nFor comparison, the ENTIRE page's text (not just the chosen container): {full_body_len} characters")
    if full_body_len > 300 and extracted_len < 20:
        print(
            "\n>>> The page clearly has real text somewhere, but almost none of it came from the chosen\n"
            "    container. That points to a container-selection bug (wrong element picked) rather than\n"
            "    the page genuinely being empty. Share this output so it can be fixed."
        )
    elif full_body_len < 300:
        print(
            "\n>>> The whole page has very little real HTML text, even outside the chosen container.\n"
            "    If the page visually looks content-rich, that content is likely delivered as images\n"
            "    or JavaScript-rendered widgets rather than crawlable HTML text — worth checking\n"
            "    view-source (not just the rendered page) in your browser to confirm."
        )
    return 0


def main(argv=None):
    args = parse_args(argv)

    if args.ai_analysis and not (os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_AUTH_TOKEN")):
        print("ERROR: --ai-analysis requires ANTHROPIC_API_KEY (or ANTHROPIC_AUTH_TOKEN) to be set "
              "in the environment.", file=sys.stderr)
        return 1

    if (args.ai_pages or args.ai_pages_file) and not args.ai_analysis:
        print("WARNING: --ai-pages/--ai-pages-file has no effect without --ai-analysis.",
              file=sys.stderr)

    google_credentials = None
    if args.search_data:
        if not args.gsc_site_url and not args.ga4_property_id:
            print("ERROR: --search-data requires at least one of --gsc-site-url or --ga4-property-id "
                  "(or use --gsc-import for Search Console data without API access).", file=sys.stderr)
            return 1
        needs_api_gsc = bool(args.gsc_site_url and not args.gsc_import)
        needs_api_ga4 = bool(args.ga4_property_id)
        if needs_api_gsc or needs_api_ga4:
            try:
                google_credentials = get_google_credentials(args.google_credentials_file, GOOGLE_SCOPES)
            except Exception as e:
                print(f"ERROR: {e}", file=sys.stderr)
                return 1

    base_netloc = normalize_netloc(urlparse(args.base_url).netloc)

    session = requests.Session()
    session.headers.update({
        "User-Agent": args.user_agent,
        "Accept-Language": "en,ja;q=0.9",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    })

    if args.debug_url:
        return debug_single_url(session, args.debug_url, args.en_prefix, args.timeout)

    print(f"Discovering pages from {args.sitemap_url} ...")
    try:
        sitemap_entries = collect_sitemap_urls(session, args.sitemap_url, args.timeout, args.exclude_sitemap)
    except Exception as e:
        print(f"ERROR: could not read sitemap: {e}", file=sys.stderr)
        return 1

    seen_urls = set()
    to_crawl = []
    for loc, lastmod in sitemap_entries:
        if normalize_netloc(urlparse(loc).netloc) != base_netloc:
            continue
        if not is_probably_html(loc):
            continue
        key = loc.rstrip("/")
        if key in seen_urls:
            continue
        seen_urls.add(key)
        to_crawl.append((loc, lastmod))

    if args.max_pages:
        to_crawl = to_crawl[:args.max_pages]

    print(f"Found {len(to_crawl)} pages to crawl.")
    if not to_crawl:
        print("No pages found — check --sitemap-url.", file=sys.stderr)
        return 1

    pages = []
    for i, (url, lastmod) in enumerate(to_crawl, 1):
        lang = classify_lang(url, args.en_prefix)
        print(f"[{i}/{len(to_crawl)}] ({lang}) {url}")
        page = build_page(session, url, lang, lastmod, args.timeout, base_netloc)
        pages.append(page)
        time.sleep(args.delay)

    print("Checking for duplicate titles / meta descriptions ...")
    dup_titles = find_duplicates(pages, lambda p: p.title)
    dup_meta = find_duplicates(pages, lambda p: p.meta_description)

    print("Pairing EN/JP page counterparts ...")
    pairs, unmatched_en, unmatched_ja = pair_pages(pages, args.en_prefix)
    print(f"  {len(pairs)} pairs matched, {len(unmatched_en)} EN pages and {len(unmatched_ja)} JP pages unmatched.")

    orphan_search_pages = []
    gsc_data, ga4_data = {}, {}

    if args.gsc_import:
        print(f"Loading Search Console export from {args.gsc_import} ...")
        try:
            gsc_data = load_gsc_export(args.gsc_import)
            print(f"  Loaded {len(gsc_data)} page rows from the export.")
        except Exception as e:
            print(f"WARNING: --gsc-import failed, continuing without it: {e}", file=sys.stderr)

    if args.search_data:
        print(f"Fetching Search Console / GA4 data (last {args.search_data_days} days) ...")
        try:
            if args.gsc_site_url and not args.gsc_import:
                gsc_data = fetch_gsc_data(google_credentials, args.gsc_site_url, args.search_data_days)
            if args.ga4_property_id:
                ga4_data = fetch_ga4_data(google_credentials, args.ga4_property_id, args.search_data_days)
        except Exception as e:
            print(f"WARNING: --search-data fetch failed, continuing without it: {e}", file=sys.stderr)

    if gsc_data or ga4_data:
        orphan_search_pages = attach_search_data(pages, gsc_data, ga4_data)
        print(f"  Matched Search Console/GA4 data to "
              f"{sum(1 for p in pages if p.gsc_impressions is not None or p.ga4_sessions is not None)} pages.")

    gsc_linking_sites, gsc_linking_text = None, None
    if args.gsc_links_import:
        print(f"Loading Search Console Links report export(s): {args.gsc_links_import} ...")
        top_linked_pages = {}
        for path in args.gsc_links_import:
            try:
                table_type, data = load_gsc_links_export(path)
            except Exception as e:
                print(f"WARNING: could not classify/load {path}, skipping: {e}", file=sys.stderr)
                continue
            print(f"  {path} -> classified as '{table_type}' ({len(data)} rows).")
            if table_type == "pages":
                top_linked_pages.update(data)
            elif table_type == "sites":
                if gsc_linking_sites is not None:
                    print(f"WARNING: multiple files classified as 'sites' — {path} overwrites "
                          f"the previous one.", file=sys.stderr)
                gsc_linking_sites = data
            elif table_type == "text":
                if gsc_linking_text is not None:
                    print(f"WARNING: multiple files classified as 'text' — {path} overwrites "
                          f"the previous one.", file=sys.stderr)
                gsc_linking_text = data
        if top_linked_pages:
            unmatched_links = attach_gsc_external_links(pages, top_linked_pages)
            print(f"  Matched external-link counts to "
                  f"{sum(1 for p in pages if p.external_backlinks is not None)} pages.")
            if unmatched_links:
                print(f"  {len(unmatched_links)} linked-page URL(s) from the export didn't match "
                      f"a crawled page.", file=sys.stderr)

    if args.skip_broken_links:
        print("Skipping broken-link check (--skip-broken-links).")
    else:
        print("Checking internal links for broken status (this can take a while) ...")
    broken_links, link_sources = check_broken_links(session, pages, args.timeout, args.delay, skip=args.skip_broken_links)

    print("Generating issue list ...")
    issues = generate_issues(pages, pairs, unmatched_en, unmatched_ja, broken_links, link_sources,
                              dup_titles, dup_meta, args.thin_words, args.thin_chars, args.en_prefix)

    ai_results = None
    if args.ai_analysis:
        keyword_map = load_keyword_map(args.ai_keyword_map) if args.ai_keyword_map else {}

        selected_urls = None
        try:
            selector_entries = parse_page_selectors(args.ai_pages, args.ai_pages_file)
        except Exception as e:
            # A bad --ai-pages-file must hard-fail rather than silently
            # falling back to an unrestricted (full-cost) AI run.
            print(f"ERROR: could not read --ai-pages-file: {e}", file=sys.stderr)
            return 1
        if selector_entries:
            selected_urls, unmatched = match_page_selectors(pages, selector_entries)
            if unmatched:
                print(f"WARNING: {len(unmatched)} --ai-pages/--ai-pages-file entry(ies) matched no "
                      f"crawled page: {unmatched}", file=sys.stderr)
            excluded_not_ok = sorted(u for u in selected_urls
                                      if not next(p for p in pages if p.url == u).ok)
            if excluded_not_ok:
                print(f"WARNING: {len(excluded_not_ok)} selected page(s) didn't crawl successfully "
                      f"and will be excluded from AI analysis: {excluded_not_ok}", file=sys.stderr)

        ai_results = run_ai_analysis(pages, args.ai_model, keyword_map, selected_urls)

    print("Writing workbook ...")
    wb = build_workbook(pages, issues, pairs, unmatched_en, unmatched_ja, broken_links, link_sources,
                         args.thin_words, args.thin_chars, args.en_prefix, ai_results, orphan_search_pages,
                         gsc_linking_sites, gsc_linking_text)
    wb.save(args.output)

    counts = defaultdict(int)
    for i in issues:
        counts[i.severity] += 1
    print(f"\nDone. Wrote {args.output}")
    print(f"Issues found — High: {counts[HIGH]}, Medium: {counts[MEDIUM]}, Low: {counts[LOW]}, Info: {counts[INFO]}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
